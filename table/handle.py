import openpyxl  # 导入模块openpyxl

#读取xlsx文件
def read():
    """ 读取xlsx文件
    """
    # 打开Excel表格
    refer_excel = openpyxl.load_workbook('table1.xlsx')
    # 获取指定Sheet表单页
    refer_sheet = refer_excel['Sheet1']
    # 创建字典：创建一个以型号为key，以品牌为value的字典
    dict = {}
    # 行循环：从第二行开始循环，到最后一行截止
    for row in range(2, refer_sheet.max_row + 1):
        # 读取cell单元格中的数据
        name = (refer_sheet.cell(row=row, column=1)).value  # 场所名称
        code = (refer_sheet.cell(row=row, column=2)).value  # 场所码
        # 以型号为key 以品牌为value
        dict[code] = name
    print("return dict: ", dict)
    return dict

if __name__ == '__main__':
    read()